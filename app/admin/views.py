import datetime
import json
import os
import time
from decimal import Decimal

from yandex_geocoder import Client
import requests
import openpyxl
import xlrd
from flask import (
    Blueprint,
    abort,
    flash,
    redirect,
    render_template,
    request,
    url_for,
)
from flask_login import current_user, login_required
from flask_rq import get_queue

from app import db
from app.admin.forms import (
    ChangeAccountTypeForm,
    ChangeUserEmailForm,
    InviteUserForm,
    NewUserForm,
)
from app.decorators import admin_required
from app.emailing import send_email
from app.models import EditableHTML, Role, User, Flat

admin = Blueprint('admin', __name__)


@admin.route('/')
@login_required
@admin_required
def index():
    """Admin dashboard page."""
    return render_template('admin/index.html')


@admin.route('/new-user', methods=['GET', 'POST'])
@login_required
@admin_required
def new_user():
    """Create a new user."""
    form = NewUserForm()
    if form.validate_on_submit():
        user = User(
            role=form.role.data,
            first_name=form.first_name.data,
            last_name=form.last_name.data,
            email=form.email.data,
            password=form.password.data)
        db.session.add(user)
        db.session.commit()
        flash('User {} successfully created'.format(user.full_name()),
              'form-success')
    return render_template('admin/new_user.html', form=form)


@admin.route('/invite-user', methods=['GET', 'POST'])
@login_required
@admin_required
def invite_user():
    """Invites a new user to create an account and set their own password."""
    form = InviteUserForm()
    if form.validate_on_submit():
        user = User(
            role=form.role.data,
            first_name=form.first_name.data,
            last_name=form.last_name.data,
            email=form.email.data)
        db.session.add(user)
        db.session.commit()
        token = user.generate_confirmation_token()
        invite_link = url_for(
            'account.join_from_invite',
            user_id=user.id,
            token=token,
            _external=True)
        get_queue().enqueue(
            send_email,
            recipient=user.email,
            subject='You Are Invited To Join',
            template='account/email/invite',
            user=user,
            invite_link=invite_link,
        )
        flash('User {} successfully invited'.format(user.full_name()),
              'form-success')
    return render_template('admin/new_user.html', form=form)


@admin.route('/users')
@login_required
@admin_required
def registered_users():
    """View all registered users."""
    users = User.query.all()
    roles = Role.query.all()
    return render_template(
        'admin/registered_users.html', users=users, roles=roles)


ALLOWED_EXTENSIONS = {'xlsx'}
UPLOAD_FOLDER = '/uploads'

def allowed_file(filename):
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

@admin.route('/updatelocation', methods=['GET', 'POST'])
@login_required
@admin_required
def updatelocation():
    r=request.form
    lng=r['lng']
    lat = r['lat']
    id=r['id']
    prop = Flat.query.filter_by(id=id).first()
    prop.prop_type=r['prop_type']
    prop.address=r['Address']
    prop.priceN = r['priceN']
    prop.plus = r['plus']
    prop.price_k = r['priceK']
    prop.market = r['market']
    prop.profit = r['profit']
    prop.percent = r['profitpercent']
    prop.comment = r['comment']

    if not lng=='':
        prop.lng=lng
    if not lat=='':
        prop.lat=lat
    db.session.add(prop)
    db.session.commit()
    return redirect(url_for('admin.flats'))


@admin.route('/getposition/<string:address>', methods=['GET', 'POST'])
@login_required
@admin_required
def getposition(address):
    try:
        client = Client("637f2780-51d5-4978-aa6b-ce5b58e4cba5")
        coordinates = client.coordinates(address)
        position ={}
        position['lat'] = float(coordinates[1])
        position['lng'] = float(coordinates[0])
        return  '{} {}'.format(position['lat'], position['lng'])
    except Exception as e:
        return 'exception'

@admin.route('/flats', methods=['GET', 'POST'])
@login_required
@admin_required
def flats():
    if request.method == 'POST':
        # check if the post request has the file part
        if len(request.files)<1:
            flash('No file part')
            return redirect(request.url)
        file = request.files['uploadedfile']
        # If the user does not select a file, the browser submits an
        # empty file without a filename.
        if file.filename == '':
            flash('No selected file')
            return redirect(request.url)
        if file and allowed_file(file.filename):
            filename = file.filename
            file.save(filename)


            #reading excel
            wookbook = openpyxl.load_workbook(filename, data_only=True)
            worksheet = wookbook.active
            Flat.query.delete()
            for i in range(3, worksheet.max_row):
                if worksheet.cell(row=i, column=1).value==None :
                    continue
                flat_to_save=Flat()
                flat_to_save.prop_type = worksheet.cell(row=i, column=1).value
                flat_to_save.address = worksheet.cell(row=i, column=2).value
                flat_to_save.Valid_date = worksheet.cell(row=i, column=4).value
                flat_to_save.priceN = worksheet.cell(row=i, column=5).value
                flat_to_save.plus = worksheet.cell(row=i, column=6).value
                flat_to_save.price_k = worksheet.cell(row=i, column=7).value
                flat_to_save.market = worksheet.cell(row=i, column=8).value
                flat_to_save.profit = worksheet.cell(row=i, column=9).value
                flat_to_save.percent = worksheet.cell(row=i, column=10).value
                flat_to_save.comment = worksheet.cell(row=i, column=12).value
                search_string=flat_to_save.address
                # search_string = search_string.split(':')[1] or search_string
                if '????,' in search_string:
                    search_string = search_string.split('????,')[1] or search_string
                    search_string = search_string.split(':')[0] or search_string
                elif '???????????????????? ??????., ' in search_string:
                    search_string = search_string.split('???????????????????? ??????., ')[1] or search_string
                    search_string = search_string.split(':')[0] or search_string
                elif '???????????????????? ??????????????,' in search_string:
                    search_string = search_string.split('???????????????????? ??????????????,')[1] or search_string
                    search_string = search_string.split(':')[0] or search_string
                try:
                    client = Client("637f2780-51d5-4978-aa6b-ce5b58e4cba5")
                    coordinates = client.coordinates(search_string)
                    flat_to_save.lat=coordinates[1]
                    flat_to_save.lng=coordinates[0]
                except Exception as e:
                    err=e
                    #rosreestr accessible only from russia
                    # kn_string=""
                    # if "??/??:" in flat_to_save.address :
                    #     kn_string = flat_to_save.address.split('??/??:')[1]
                    #     kn_string = kn_string.split(' ')[1]
                    # elif "??/??" in flat_to_save.address :
                    #     kn_string = flat_to_save.address.split('??/??')[1]
                    #     kn_string = kn_string.split(' ')[1]
                    # elif "?????????????????????? ??????????" in flat_to_save.address :
                    #     kn_string = flat_to_save.address.split('?????????????????????? ??????????')[1]
                    #     kn_string = kn_string.split(' ')[1]
                    #
                    # if kn_string != "":
                    #     kn_string=kn_string.replace(";","")
                    #     kn_string = kn_string.replace(",", "")
                    #     headers = {
                    #         'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_11_5) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/50.0.2661.102 Safari/537.36'}
                    #
                    #     resp = requests.get('https://rosreestr.gov.ru/api/online/fir_objects/'+kn_string, verify=False, headers=headers)
                    #     # while resp.status_code == 204:
                    #     #     time.sleep(1)
                    #     #     resp = requests.get('https://rosreestr.gov.ru/api/online/fir_objects/' + kn_string,
                    #     #                         verify=False, headers=headers)
                    #
                    #     if resp.status_code==200:
                    #         try:
                    #             result = resp.json()
                    #             if len(result)>0:
                    #
                    #                 addr=result[0]['addressNotes']
                    #                 try:
                    #                     client = Client("637f2780-51d5-4978-aa6b-ce5b58e4cba5")
                    #                     coordinates = client.coordinates(addr)
                    #                     flat_to_save.lat = coordinates[1]
                    #                     flat_to_save.lng = coordinates[0]
                    #                 except Exception as e:
                    #                     errormessage=e
                    #         except Exception as e:
                    #             errormessage = e
                    #         # elif "??/??" in flat_to_save.address :
                    #         #     kn_string = flat_to_save.address.split('??/??')[1]
                    #         #     kn=search_string = kn_string.split(' ')[1]
                    #         #     headers = {
                    #         #         'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_11_5) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/50.0.2661.102 Safari/537.36'}
                    #         #
                    #         #     resp = requests.get('https://rosreestr.gov.ru/api/online/fir_objects/'+kn, verify=False, headers=headers)
                    #         #     result= resp.json()
                    #         #     if len(result)>0:
                    #         #         addr=result[0]['addressNotes']
                    #         #         try:
                    #         #             client = Client("637f2780-51d5-4978-aa6b-ce5b58e4cba5")
                    #         #             coordinates = client.coordinates(addr)
                    #         #             flat_to_save.lat = coordinates[1]
                    #         #             flat_to_save.lng = coordinates[0]
                    #         #         except Exception as e:
                    #         #             errormessage=e





                #google version
                # search_string = search_string.replace(".", "")
                # search_string = search_string.replace(",", "")
                # #search_string = search_string.replace(":", " ")
                # search_string = search_string.replace(" ", "%20")
                # url = "https://maps.googleapis.com/maps/api/place/textsearch/json?query="+search_string+"&key="+"AIzaSyB1-XsaWMPoip4mCid-NSbDmb_kpe0CI_4"
                # payload = {}
                # headers = {}
                # response = requests.request("GET", url, headers=headers, data=payload)
                # rt=response.json()
                # if len(rt['results'])>0:
                #     result=rt['results'][0]
                #     flat_to_save.lng=result['geometry']['location']['lng']
                #     flat_to_save.lat=result['geometry']['location']['lat']

                flat_to_save.add()

            os.remove(filename)

    flats= Flat.query.all()
    total_deparments=len(flats)
    departments_with_location=0
    for f in flats:
        if f.lat and f.lng:
            f.loc=True
            departments_with_location=departments_with_location+1
        # f.Valid_date=datetime.datetime.strptime(f.Valid_date)
    res=[]
    for f in flats:
        res.append(f.to_dictionary())
    json_string = json.dumps(res, ensure_ascii=False)
    return render_template(
        'admin/flats.html', flats=flats,flats_json=json_string,total_deparments=total_deparments,departments_with_location=departments_with_location)




@admin.route('/user/<int:user_id>')
@admin.route('/user/<int:user_id>/info')
@login_required
@admin_required
def user_info(user_id):
    """View a user's profile."""
    user = User.query.filter_by(id=user_id).first()
    if user is None:
        abort(404)
    return render_template('admin/manage_user.html', user=user)


@admin.route('/user/<int:user_id>/change-email', methods=['GET', 'POST'])
@login_required
@admin_required
def change_user_email(user_id):
    """Change a user's email."""
    user = User.query.filter_by(id=user_id).first()
    if user is None:
        abort(404)
    form = ChangeUserEmailForm()
    if form.validate_on_submit():
        user.email = form.email.data
        db.session.add(user)
        db.session.commit()
        flash('Email for user {} successfully changed to {}.'.format(
            user.full_name(), user.email), 'form-success')
    return render_template('admin/manage_user.html', user=user, form=form)


@admin.route(
    '/user/<int:user_id>/change-account-type', methods=['GET', 'POST'])
@login_required
@admin_required
def change_account_type(user_id):
    """Change a user's account type."""
    if current_user.id == user_id:
        flash('You cannot change the type of your own account. Please ask '
              'another administrator to do this.', 'error')
        return redirect(url_for('admin.user_info', user_id=user_id))

    user = User.query.get(user_id)
    if user is None:
        abort(404)
    form = ChangeAccountTypeForm()
    if form.validate_on_submit():
        user.role = form.role.data
        db.session.add(user)
        db.session.commit()
        flash('Role for user {} successfully changed to {}.'.format(
            user.full_name(), user.role.name), 'form-success')
    return render_template('admin/manage_user.html', user=user, form=form)


@admin.route('/user/<int:user_id>/delete')
@login_required
@admin_required
def delete_user_request(user_id):
    """Request deletion of a user's account."""
    user = User.query.filter_by(id=user_id).first()
    if user is None:
        abort(404)
    return render_template('admin/manage_user.html', user=user)


@admin.route('/user/<int:user_id>/_delete')
@login_required
@admin_required
def delete_user(user_id):
    """Delete a user's account."""
    if current_user.id == user_id:
        flash('You cannot delete your own account. Please ask another '
              'administrator to do this.', 'error')
    else:
        user = User.query.filter_by(id=user_id).first()
        db.session.delete(user)
        db.session.commit()
        flash('Successfully deleted user %s.' % user.full_name(), 'success')
    return redirect(url_for('admin.registered_users'))


@admin.route('/_update_editor_contents', methods=['POST'])
@login_required
@admin_required
def update_editor_contents():
    """Update the contents of an editor."""

    edit_data = request.form.get('edit_data')
    editor_name = request.form.get('editor_name')

    editor_contents = EditableHTML.query.filter_by(
        editor_name=editor_name).first()
    if editor_contents is None:
        editor_contents = EditableHTML(editor_name=editor_name)
    editor_contents.value = edit_data

    db.session.add(editor_contents)
    db.session.commit()

    return 'OK', 200
